
"""
PHD V2.2 Excel parser.

Reads the PINC AI Premier Healthcare Database Data Dictionary Excel file
and returns structured Python objects for ingestion into Delta.

PHD V2.2 workbook layout (case-insensitive sheet/column matching):
  - Sheet "Table Descriptions" (or "Tables", "Table Desc"):
      TABLE_NAME | DESCRIPTION | GRAIN | PRIMARY_JOIN_KEY | IS_ADDON
  - Sheet "Column Descriptions" (or "Data Dictionary", "Columns", the first large sheet):
      TABLE_NAME | COLUMN_NAME | DATA_TYPE | DESCRIPTION | VALID_VALUES
      | IS_PRIMARY_KEY | IS_FOREIGN_KEY | IS_NULLABLE

If the workbook has no dedicated Tables sheet, table metadata is derived
from the unique TABLE_NAME values in the columns sheet.

File-lock note: on Windows, if Excel has the file open, Python gets
PermissionError.  The caller should copy the file to a temp path first.
"""
from __future__ import annotations

import logging
import re
import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── Add-on tables require extra Premier licensing ──────────────────────────────
_ADDON_TABLES: frozenset[str] = frozenset({
    "genlab", "vitals", "lab_res", "lab_sens",
    "proc_supply", "mortality", "tokens",
    "mother_infant_link", "pat_sdoh",
})

# Tables that join on medrec_key instead of pat_key
_MEDREC_JOIN_TABLES: frozenset[str] = frozenset({"mortality", "tokens"})

# ── Sheet name aliases (lower-cased) ──────────────────────────────────────────
_TABLE_SHEET_ALIASES = (
    "table descriptions", "table desc", "tables", "table_descriptions",
    "table list", "tablelist",
)
_COLUMN_SHEET_ALIASES = (
    "column descriptions", "column desc", "columns", "data dictionary",
    "data_dictionary", "phd data dictionary", "field descriptions",
    "field desc", "fields",
)

# ── Column header aliases (lower-cased, stripped) ─────────────────────────────
_COL_MAP: dict[str, list[str]] = {
    "table_name":      ["table name", "table_name", "tablename", "tbl name", "tbl_name"],
    "column_name":     ["column name", "column_name", "columnname", "field name",
                        "field_name", "col name", "col_name", "variable name"],
    "data_type":       ["data type", "data_type", "datatype", "type", "field type"],
    "description":     ["description", "column description", "field description",
                        "col description", "definition", "desc"],
    "valid_values":    ["valid values", "valid_values", "validvalues", "allowed values",
                        "sample values", "values", "enumerated values", "example values"],
    "is_primary_key":  ["primary key", "primary_key", "pk", "is_primary_key",
                        "is primary key", "key"],
    "is_foreign_key":  ["foreign key", "foreign_key", "fk", "is_foreign_key",
                        "is foreign key"],
    "is_nullable":     ["nullable", "is_nullable", "null allowed", "null", "nulls"],
    "code_set_type":   ["code set", "code_set", "code set type", "codeset",
                        "code type", "icd type"],
}

_TABLE_COL_MAP: dict[str, list[str]] = {
    "table_name":      ["table name", "table_name", "tablename", "tbl name"],
    "description":     ["description", "table description", "definition", "desc"],
    "grain":           ["grain", "level", "granularity", "unit of analysis"],
    "primary_join_key": ["primary key", "join key", "primary join key", "pk column"],
    "is_addon":        ["add-on", "addon", "is add-on", "add on", "is_addon",
                        "licensed", "requires license"],
}


# ── Output dataclasses ─────────────────────────────────────────────────────────

@dataclass
class ParsedTableRow:
    table_name: str
    description: str = ""
    grain: str = ""
    primary_join_key: str = "pat_key"
    is_addon: bool = False
    use_cases: list[str] = field(default_factory=list)


@dataclass
class ParsedColumnRow:
    table_name: str
    column_name: str
    data_type: str = ""
    description: str = ""
    valid_values: str = ""
    is_primary_key: bool = False
    is_foreign_key: bool = False
    is_nullable: bool = True
    code_set_type: str = ""
    embedding_text: str = ""


@dataclass
class DiscoveredRelationship:
    """
    A FK relationship found by scanning hyperlinks or text in the Excel.
    Example: PATDEMO.prov_id → PROVIDERS (from a hyperlink on that cell).
    These are merged with the hardcoded seed in the ingestor.
    """
    from_table: str
    from_column: str
    to_table: str
    evidence: str  # "hyperlink" | "text_reference" | "column_reference"


@dataclass
class ParsedWorkbook:
    tables: list[ParsedTableRow]
    columns: list[ParsedColumnRow]
    discovered_relationships: list[DiscoveredRelationship] = field(default_factory=list)
    source_file_name: str = ""
    sheet_names_found: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ── Parser ─────────────────────────────────────────────────────────────────────

class PhDExcelParser:
    """
    Reads a PHD V2.2 Excel file using openpyxl and returns a ParsedWorkbook.

    Handles:
    - Both flat (single large sheet) and split (Tables + Columns sheets) layouts
    - Case-insensitive header matching via _COL_MAP / _TABLE_COL_MAP
    - Windows file locks (caller should pass a temp-copy path)
    - Missing / blank rows — skipped silently
    - Derives table metadata from column sheet when no Tables sheet exists
    """

    def parse(self, file_path: str | Path, source_file_name: str = "") -> ParsedWorkbook:
        import openpyxl

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Data dictionary file not found: {path}")

        logger.info("Opening PHD Excel: %s", path)
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot open {path} — file may be locked by Excel. "
                "Close the file in Excel and retry, or the caller should copy it to a temp path first."
            ) from exc

        sheet_names = wb.sheetnames
        logger.info("Sheets found: %s", sheet_names)

        table_sheet = self._find_sheet(wb, sheet_names, _TABLE_SHEET_ALIASES)
        col_sheet = self._find_sheet(wb, sheet_names, _COLUMN_SHEET_ALIASES)

        # Fallback: if no column sheet found, use the largest sheet
        if col_sheet is None:
            col_sheet = max(
                (wb[s] for s in sheet_names),
                key=lambda ws: ws.max_row,
            )
            logger.warning(
                "No recognized column sheet found. Using largest sheet: %s",
                col_sheet.title,
            )

        warnings: list[str] = []
        known_sheet_names: frozenset[str] = frozenset(s.lower() for s in sheet_names)

        tables = self._parse_table_sheet(table_sheet, warnings) if table_sheet else []
        columns, discovered_rels = self._parse_column_sheet(
            col_sheet, warnings, known_sheet_names
        )

        # Derive table rows from column data when no table sheet exists
        if not tables and columns:
            tables = self._derive_tables_from_columns(columns)
            warnings.append(
                "No Tables sheet found — table metadata derived from column data. "
                "Table descriptions will be empty."
            )

        # Post-process: set embedding_text
        for col in columns:
            col.embedding_text = _build_embedding_text(col)

        # Post-process: set add-on flag from canonical set
        for tbl in tables:
            if tbl.table_name.lower() in _ADDON_TABLES:
                tbl.is_addon = True
            if tbl.table_name.lower() in _MEDREC_JOIN_TABLES:
                tbl.primary_join_key = "medrec_key"

        logger.info(
            "Parsed %d tables, %d columns, %d discovered relationships. Warnings: %d",
            len(tables), len(columns), len(discovered_rels), len(warnings),
        )
        return ParsedWorkbook(
            tables=tables,
            columns=columns,
            discovered_relationships=discovered_rels,
            source_file_name=source_file_name or path.name,
            sheet_names_found=sheet_names,
            warnings=warnings,
        )

    # ── Sheet finders ──────────────────────────────────────────────────────────

    @staticmethod
    def _find_sheet(wb, sheet_names: list[str], aliases: tuple[str, ...]):
        for name in sheet_names:
            if name.lower().strip() in aliases:
                return wb[name]
        return None

    # ── Table sheet parser ─────────────────────────────────────────────────────

    def _parse_table_sheet(self, ws, warnings: list[str]) -> list[ParsedTableRow]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header_map = _build_header_map(rows[0], _TABLE_COL_MAP)
        if "table_name" not in header_map:
            warnings.append(
                f"Sheet '{ws.title}' has no recognisable TABLE_NAME column — skipped."
            )
            return []

        results: list[ParsedTableRow] = []
        for raw_row in rows[1:]:
            row = dict(zip(range(len(raw_row)), raw_row))
            tbl_name = _cell(row, header_map.get("table_name"))
            if not tbl_name:
                continue
            results.append(ParsedTableRow(
                table_name=tbl_name.lower().strip(),
                description=_cell(row, header_map.get("description")),
                grain=_cell(row, header_map.get("grain")),
                primary_join_key=_cell(row, header_map.get("primary_join_key")) or "pat_key",
                is_addon=_bool_cell(row, header_map.get("is_addon")),
            ))
        return results

    # ── Column sheet parser ────────────────────────────────────────────────────

    def _parse_column_sheet(
        self,
        ws,
        warnings: list[str],
        known_sheet_names: frozenset[str] = frozenset(),
    ) -> tuple[list[ParsedColumnRow], list[DiscoveredRelationship]]:
        """
        Parse column rows AND extract lookup-table relationships from:
          1. Cell hyperlinks  (#PROVIDERS!A1  →  to_table = "providers")
          2. Text patterns in valid_values/description ("See PROVIDERS", "PROVIDERS")
          3. A dedicated "Lookup Table" / "Reference Table" column

        Uses iter_rows(values_only=False) so we can read both .value and .hyperlink.
        """
        # Collect cell rows (objects, not values) for hyperlink access
        all_cell_rows = list(ws.iter_rows(values_only=False))
        if not all_cell_rows:
            return [], []

        # Convert to value rows for header matching
        value_rows = [tuple(c.value for c in row) for row in all_cell_rows]

        # Find header row
        header_row_idx = 0
        header_map: dict[str, int] = {}
        for idx, row in enumerate(value_rows[:10]):
            candidate = _build_header_map(row, _COL_MAP)
            if "table_name" in candidate and "column_name" in candidate:
                header_row_idx = idx
                header_map = candidate
                break

        # Also look for a lookup/reference column
        ref_col_idx = _find_reference_column(value_rows[header_row_idx] if value_rows else ())

        if not header_map:
            warnings.append(
                f"Sheet '{ws.title}' — could not identify TABLE_NAME + COLUMN_NAME headers. "
                "First 10 rows examined."
            )
            return [], []

        results: list[ParsedColumnRow] = []
        discovered: list[DiscoveredRelationship] = []

        for cell_row, raw_row in zip(
            all_cell_rows[header_row_idx + 1:],
            value_rows[header_row_idx + 1:],
        ):
            row = dict(zip(range(len(raw_row)), raw_row))
            tbl_name = _cell(row, header_map.get("table_name"))
            col_name = _cell(row, header_map.get("column_name"))
            if not tbl_name or not col_name:
                continue

            tbl_name = tbl_name.lower().strip()
            col_name = col_name.lower().strip()

            parsed_col = ParsedColumnRow(
                table_name=tbl_name,
                column_name=col_name,
                data_type=_cell(row, header_map.get("data_type")),
                description=_cell(row, header_map.get("description")),
                valid_values=_cell(row, header_map.get("valid_values")),
                is_primary_key=_bool_cell(row, header_map.get("is_primary_key")),
                is_foreign_key=_bool_cell(row, header_map.get("is_foreign_key")),
                is_nullable=_bool_cell(row, header_map.get("is_nullable"), default=True),
                code_set_type=_cell(row, header_map.get("code_set_type")),
            )
            results.append(parsed_col)

            # ── Hyperlink scanning ─────────────────────────────────────────────
            # Check every cell in the row for a hyperlink pointing to a sheet
            for cell in cell_row:
                ref_table = _extract_sheet_ref_from_hyperlink(cell, known_sheet_names)
                if ref_table and ref_table != tbl_name:
                    discovered.append(DiscoveredRelationship(
                        from_table=tbl_name,
                        from_column=col_name,
                        to_table=ref_table,
                        evidence="hyperlink",
                    ))

            # ── Reference column scanning ──────────────────────────────────────
            if ref_col_idx is not None:
                ref_val = _cell(row, ref_col_idx).lower().strip()
                if ref_val and ref_val in known_sheet_names and ref_val != tbl_name:
                    discovered.append(DiscoveredRelationship(
                        from_table=tbl_name,
                        from_column=col_name,
                        to_table=ref_val,
                        evidence="column_reference",
                    ))

            # ── Text pattern scanning in valid_values / description ────────────
            for field_key in ("valid_values", "description"):
                text = _cell(row, header_map.get(field_key))
                ref_table = _extract_table_ref_from_text(text, known_sheet_names, tbl_name)
                if ref_table:
                    discovered.append(DiscoveredRelationship(
                        from_table=tbl_name,
                        from_column=col_name,
                        to_table=ref_table,
                        evidence="text_reference",
                    ))

        # Deduplicate discovered relationships
        seen: set[tuple[str, str, str]] = set()
        unique_discovered: list[DiscoveredRelationship] = []
        for r in discovered:
            key = (r.from_table, r.from_column, r.to_table)
            if key not in seen:
                seen.add(key)
                unique_discovered.append(r)

        if unique_discovered:
            logger.info(
                "Discovered %d FK relationships from Excel hyperlinks/text",
                len(unique_discovered),
            )

        return results, unique_discovered

    # ── Derive table rows from column data ─────────────────────────────────────

    @staticmethod
    def _derive_tables_from_columns(columns: list[ParsedColumnRow]) -> list[ParsedTableRow]:
        seen: dict[str, ParsedTableRow] = {}
        for col in columns:
            t = col.table_name
            if t not in seen:
                seen[t] = ParsedTableRow(
                    table_name=t,
                    primary_join_key=(
                        "medrec_key" if t in _MEDREC_JOIN_TABLES else "pat_key"
                    ),
                    is_addon=t in _ADDON_TABLES,
                )
        return list(seen.values())


# ── Helpers ────────────────────────────────────────────────────────────────────

def _build_header_map(header_row: tuple, col_map: dict[str, list[str]]) -> dict[str, int]:
    """Return {field_name: column_index} from a header row."""
    result: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        normalised = str(cell).lower().strip()
        for field_name, aliases in col_map.items():
            if normalised in aliases and field_name not in result:
                result[field_name] = idx
                break
    return result


def _cell(row: dict[int, Any], col_idx: int | None) -> str:
    if col_idx is None:
        return ""
    val = row.get(col_idx)
    if val is None:
        return ""
    return str(val).strip()


def _bool_cell(row: dict[int, Any], col_idx: int | None, default: bool = False) -> bool:
    if col_idx is None:
        return default
    val = row.get(col_idx)
    if val is None:
        return default
    s = str(val).lower().strip()
    return s in ("y", "yes", "true", "1", "x", "pk", "fk", "key")


def _build_embedding_text(col: ParsedColumnRow) -> str:
    """
    Concatenate searchable text for this column.
    Phase 12: this is the field that gets embedded into Vector Search.
    """
    parts = [col.column_name, col.description, col.valid_values or ""]
    return " ".join(p for p in parts if p).strip()


# ── Lookup table relationship helpers ─────────────────────────────────────────

# Known reference/lookup column header names (lower-cased)
_REF_COL_ALIASES = (
    "lookup table", "lookup_table", "reference table", "reference_table",
    "refers to", "ref table", "fk table", "foreign table", "links to",
    "see table", "lookup", "reference",
)

# Text patterns that signal a lookup table reference
# "See PROVIDERS", "See PROVIDERS table", "PROVIDERS lookup"
_SEE_PATTERN = re.compile(
    r'\bsee\s+([A-Z_][A-Z0-9_]+)(?:\s+table)?\b'
    r'|\b([A-Z_][A-Z0-9_]+)\s+(?:lookup|table|reference)\b',
    re.IGNORECASE,
)


def _find_reference_column(header_row: tuple) -> int | None:
    """Return column index of a dedicated lookup/reference column, or None."""
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        if str(cell).lower().strip() in _REF_COL_ALIASES:
            return idx
    return None


def _extract_sheet_ref_from_hyperlink(cell, known_sheet_names: frozenset[str]) -> str | None:
    """
    If the cell has an internal hyperlink (#SHEETNAME!A1), return the
    lower-cased sheet name if it matches a known sheet.  Returns None otherwise.
    """
    if cell.hyperlink is None:
        return None
    target = getattr(cell.hyperlink, "target", None) or ""
    if not target.startswith("#"):
        return None
    # Strip leading # and trailing !CellRef  →  "PROVIDERS"
    sheet_part = target[1:].split("!")[0].strip().lower()
    if sheet_part in known_sheet_names:
        return sheet_part
    return None


def _extract_table_ref_from_text(
    text: str,
    known_sheet_names: frozenset[str],
    current_table: str,
) -> str | None:
    """
    Scan free text (valid_values, description) for patterns like
    "See PROVIDERS" or "PROVIDERS table".  Returns the lower-cased table
    name if it matches a known sheet and is not the current table.
    """
    if not text:
        return None
    for match in _SEE_PATTERN.finditer(text):
        candidate = (match.group(1) or match.group(2) or "").lower().strip()
        if candidate and candidate in known_sheet_names and candidate != current_table:
            return candidate
    return None


def safe_open_excel(source_path: str | Path) -> Path:
    """
    Copy the Excel file to a temp location and return the temp path.
    Avoids PermissionError when the file is open in Excel on Windows.
    """
    source = Path(source_path)
    tmp = tempfile.NamedTemporaryFile(
        suffix=source.suffix, delete=False, prefix="phd_dict_"
    )
    tmp.close()
    shutil.copy2(source, tmp.name)
    logger.debug("Copied Excel to temp path: %s", tmp.name)
    return Path(tmp.name)
