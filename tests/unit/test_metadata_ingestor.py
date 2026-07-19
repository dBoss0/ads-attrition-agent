"""
Phase 4 unit tests — Metadata Ingestor (no Spark, no real Excel needed).

Tests cover:
  - ExcelParser header mapping (both layouts: flat and split)
  - Embedding text generation
  - Add-on table flag and join key auto-detection
  - Relationship seed: join key constraints
  - Business rules seed: all mandatory rules present
  - Ingestor: validation logic, lock contention, bytes ingestion path
"""
from __future__ import annotations

import os
import tempfile
from unittest.mock import MagicMock, patch

import pytest
import openpyxl


# ── Helpers — build minimal test Excel workbooks ───────────────────────────────

def _make_flat_workbook(rows: list[list]) -> str:
    """Write a single-sheet Excel file and return its tmp path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"
    for row in rows:
        ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _make_split_workbook(
    table_rows: list[list],
    col_rows: list[list],
) -> str:
    """Write a two-sheet Excel and return its tmp path."""
    wb = openpyxl.Workbook()
    ws_tables = wb.active
    ws_tables.title = "Table Descriptions"
    for row in table_rows:
        ws_tables.append(row)
    ws_cols = wb.create_sheet("Column Descriptions")
    for row in col_rows:
        ws_cols.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ── ExcelParser: header mapping ────────────────────────────────────────────────

class TestPhDExcelParserFlat:
    """Tests for single-sheet 'flat' layout."""

    def test_parses_basic_columns(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description", "Valid Values"],
            ["patdemo", "pat_key", "STRING", "Encounter key", ""],
            ["patdemo", "disc_mon", "STRING", "Discharge month YYYYMM", ""],
            ["paticd_diag", "icd_code", "STRING", "ICD code", ""],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path, "test.xlsx")
            assert len(parsed.columns) == 3
            assert parsed.columns[0].table_name == "patdemo"
            assert parsed.columns[0].column_name == "pat_key"
            assert parsed.columns[0].data_type == "STRING"
        finally:
            os.unlink(path)

    def test_derives_tables_when_no_table_sheet(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description"],
            ["patdemo", "pat_key", "STRING", "Encounter key"],
            ["patdemo", "disc_mon", "STRING", "Discharge month"],
            ["paticd_diag", "icd_code", "STRING", "ICD code"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            table_names = {t.table_name for t in parsed.tables}
            assert "patdemo" in table_names
            assert "paticd_diag" in table_names
        finally:
            os.unlink(path)

    def test_addon_flag_set_for_genlab(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description"],
            ["genlab", "lab_test_name", "STRING", "Lab test name"],
            ["genlab", "result_value", "STRING", "Result"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            genlab_table = next(t for t in parsed.tables if t.table_name == "genlab")
            assert genlab_table.is_addon is True
        finally:
            os.unlink(path)

    def test_mortality_uses_medrec_key(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description"],
            ["mortality", "medrec_key", "STRING", "Patient key"],
            ["mortality", "death_date", "STRING", "Date of death"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            mort_table = next(t for t in parsed.tables if t.table_name == "mortality")
            assert mort_table.primary_join_key == "medrec_key"
        finally:
            os.unlink(path)

    def test_embedding_text_built_correctly(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description", "Valid Values"],
            ["patdemo", "i_o_ind", "STRING", "Inpatient/Outpatient indicator", "I=Inpatient; O=Outpatient"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            col = parsed.columns[0]
            assert "i_o_ind" in col.embedding_text
            assert "Inpatient/Outpatient" in col.embedding_text
            assert "I=Inpatient" in col.embedding_text
        finally:
            os.unlink(path)

    def test_case_insensitive_header_matching(self):
        path = _make_flat_workbook([
            ["TABLE NAME", "COLUMN NAME", "DATA TYPE", "DESCRIPTION"],
            ["patdemo", "pat_key", "STRING", "Encounter key"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            assert len(parsed.columns) == 1
        finally:
            os.unlink(path)

    def test_skips_blank_rows(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description"],
            ["patdemo", "pat_key", "STRING", "Encounter key"],
            [None, None, None, None],
            ["patdemo", "disc_mon", "STRING", "Discharge month"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            assert len(parsed.columns) == 2
        finally:
            os.unlink(path)

    def test_boolean_pk_parsing(self):
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description", "Primary Key"],
            ["patdemo", "pat_key", "STRING", "Encounter key", "Y"],
            ["patdemo", "disc_mon", "STRING", "Discharge month", "N"],
        ])
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            assert parsed.columns[0].is_primary_key is True
            assert parsed.columns[1].is_primary_key is False
        finally:
            os.unlink(path)


class TestPhDExcelParserSplit:
    """Tests for two-sheet 'split' layout."""

    def test_reads_table_sheet_descriptions(self):
        path = _make_split_workbook(
            table_rows=[
                ["Table Name", "Description", "Grain"],
                ["patdemo", "Patient demographics", "One row per encounter"],
            ],
            col_rows=[
                ["Table Name", "Column Name", "Data Type", "Description"],
                ["patdemo", "pat_key", "STRING", "Encounter key"],
            ],
        )
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            tbl = next(t for t in parsed.tables if t.table_name == "patdemo")
            assert "demographics" in tbl.description.lower()
            assert "encounter" in tbl.grain.lower()
        finally:
            os.unlink(path)

    def test_columns_from_column_sheet(self):
        path = _make_split_workbook(
            table_rows=[
                ["Table Name", "Description"],
                ["patdemo", "Patient demographics"],
            ],
            col_rows=[
                ["Table Name", "Column Name", "Data Type", "Description"],
                ["patdemo", "pat_key", "STRING", "Encounter key"],
                ["patdemo", "disc_mon", "STRING", "Discharge month"],
            ],
        )
        try:
            from application.metadata.excel_parser import PhDExcelParser
            parsed = PhDExcelParser().parse(path)
            assert len(parsed.columns) == 2
        finally:
            os.unlink(path)


# ── Relationship seed ──────────────────────────────────────────────────────────

class TestRelationshipSeed:
    def test_returns_relationships(self):
        from application.metadata.relationship_seed import get_premier_relationships
        rels = get_premier_relationships()
        assert len(rels) > 10

    def test_icd_join_includes_version(self):
        from application.metadata.relationship_seed import get_premier_relationships
        rels = get_premier_relationships()
        icd_rel = next(r for r in rels if r.relationship_id == "rel_paticd_diag_icdcode")
        assert "icd_version" in icd_rel.join_condition

    def test_mortality_uses_medrec_key(self):
        from application.metadata.relationship_seed import get_premier_relationships
        rels = get_premier_relationships()
        mort = next(r for r in rels if r.relationship_id == "rel_patdemo_mortality")
        assert "medrec_key" in mort.join_condition

    def test_prov_enrollment_has_three_columns(self):
        from application.metadata.relationship_seed import get_premier_relationships
        rels = get_premier_relationships()
        prov = next(r for r in rels if r.relationship_id == "rel_patdemo_prov_enrollment")
        assert "prov_id" in prov.join_condition
        assert "disc_mon" in prov.join_condition
        assert "i_o_ind" in prov.join_condition

    def test_all_relationships_have_unique_ids(self):
        from application.metadata.relationship_seed import get_premier_relationships
        rels = get_premier_relationships()
        ids = [r.relationship_id for r in rels]
        assert len(ids) == len(set(ids))


# ── Business rules seed ────────────────────────────────────────────────────────

class TestBusinessRulesSeed:
    def test_returns_rules(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        assert len(rules) >= 8

    def test_icd_version_rule_exists(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        rule = next((r for r in rules if r.rule_id == "rule_icd_version_required"), None)
        assert rule is not None
        assert "icd_version" in rule.sql_pattern

    def test_inpatient_filter_rule_exists(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        rule = next((r for r in rules if r.rule_id == "rule_inpatient_filter"), None)
        assert rule is not None
        assert "i_o_ind" in rule.sql_pattern

    def test_row_number_rule_exists(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        rule = next((r for r in rules if r.rule_id == "rule_row_number_not_rank"), None)
        assert rule is not None
        assert "ROW_NUMBER" in rule.sql_pattern

    def test_all_rules_have_unique_ids(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        ids = [r.rule_id for r in rules]
        assert len(ids) == len(set(ids))

    def test_addon_rule_covers_genlab(self):
        from application.metadata.business_rules_seed import get_premier_business_rules
        rules = get_premier_business_rules()
        rule = next(r for r in rules if r.rule_id == "rule_addon_license")
        assert "genlab" in rule.applicable_tables


# ── Ingestor validation ────────────────────────────────────────────────────────

class TestMetadataIngestorValidation:
    def test_validation_fails_when_no_columns(self):
        from application.metadata.ingestor import MetadataIngestor
        from application.metadata.excel_parser import ParsedWorkbook, ParsedTableRow

        ingestor = MetadataIngestor.__new__(MetadataIngestor)
        parsed = ParsedWorkbook(tables=[], columns=[], source_file_name="test.xlsx")
        errors = MetadataIngestor._validate(parsed)
        assert any("No columns" in e for e in errors)

    def test_validation_fails_when_too_few_columns(self):
        from application.metadata.ingestor import MetadataIngestor
        from application.metadata.excel_parser import ParsedWorkbook, ParsedColumnRow

        ingestor = MetadataIngestor.__new__(MetadataIngestor)
        cols = [
            ParsedColumnRow(table_name="patdemo", column_name=f"col_{i}")
            for i in range(5)
        ]
        from application.metadata.excel_parser import ParsedTableRow
        tables = [ParsedTableRow(table_name="patdemo")]
        parsed = ParsedWorkbook(tables=tables, columns=cols)
        errors = MetadataIngestor._validate(parsed)
        assert any("Only 5 columns" in e for e in errors)

    def test_validation_passes_with_sufficient_data(self):
        from application.metadata.ingestor import MetadataIngestor
        from application.metadata.excel_parser import ParsedWorkbook, ParsedColumnRow, ParsedTableRow

        cols = [
            ParsedColumnRow(table_name="patdemo", column_name=f"col_{i}")
            for i in range(50)
        ]
        tables = [ParsedTableRow(table_name="patdemo")]
        parsed = ParsedWorkbook(tables=tables, columns=cols)
        errors = MetadataIngestor._validate(parsed)
        assert errors == []

    def test_ingest_from_bytes_uses_temp_file(self):
        """ingest_from_bytes writes to temp, parses, then cleans up."""
        from application.metadata.ingestor import MetadataIngestor

        # Build a minimal valid Excel as bytes
        path = _make_flat_workbook([
            ["Table Name", "Column Name", "Data Type", "Description"],
            *[["patdemo", f"col_{i}", "STRING", f"desc {i}"] for i in range(50)],
        ])
        try:
            with open(path, "rb") as f:
                content = f.read()
        finally:
            os.unlink(path)

        mock_spark = MagicMock()
        mock_spark.sql.return_value = MagicMock(collect=lambda: [])

        ingestor = MetadataIngestor(mock_spark)
        summary = ingestor.ingest_from_bytes(content, "test_dict.xlsx", "deepak@mu-sigma.com")

        # Spark.sql was called — ingestion ran (even if Delta writes are mocked)
        assert mock_spark.sql.called

    def test_concurrent_ingestion_rejected(self):
        """Second concurrent call returns an error immediately."""
        import threading
        from application.metadata.ingestor import MetadataIngestor, _INGEST_LOCK

        # Acquire the lock to simulate an in-progress ingestion
        _INGEST_LOCK.acquire()
        try:
            mock_spark = MagicMock()
            ingestor = MetadataIngestor(mock_spark)
            summary = ingestor.ingest_from_bytes(b"", "test.xlsx", "test")
            assert not summary.succeeded
            assert any("already in progress" in e for e in summary.errors)
        finally:
            _INGEST_LOCK.release()
